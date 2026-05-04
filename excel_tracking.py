"""
从 DHL 订单 Excel 中读取转单号列（默认识别 Sheet1 的 C 列，与业务表头一致）。
"""

from __future__ import annotations

import io
import re
from typing import BinaryIO

# 表头名命中时跳过该行（避免把「转单号」当运单）
_HEADER_HINTS = frozenset(
    {
        "转单号",
        "tracking",
        "tracking no",
        "tracking no.",
        "tracking number",
        "waybill",
        "运单号",
    }
)


def _normalize_cell(value: object) -> str | None:
    """单元格转为运单字符串；Excel 长数字常为 float。"""
    if value is None:
        return None
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip() or None
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    return s or None


def read_waybills_from_xlsx(
    path_or_file: str | BinaryIO | bytes,
    *,
    sheet_index: int = 0,
    column_index: int = 2,
    skip_header_row: bool = True,
) -> list[str]:
    """
    读取 xlsx 指定列作为转单号列表。

    sheet_index: 工作表索引，默认 0（第一张表，一般为订单明细）。
    column_index: 列索引 0=A, 1=B, 2=C（转单号），默认 2。
    skip_header_row: 若首行该列像表头则跳过。
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("读取 Excel 需要 openpyxl：pip install openpyxl") from e

    if isinstance(path_or_file, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(path_or_file), read_only=True, data_only=True)
    elif hasattr(path_or_file, "read"):
        wb = openpyxl.load_workbook(path_or_file, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(path_or_file, read_only=True, data_only=True)

    try:
        ws = wb.worksheets[sheet_index]
        out: list[str] = []
        for idx, row in enumerate(
            ws.iter_rows(min_row=1, max_col=column_index + 1, values_only=True)
        ):
            if row is None or len(row) <= column_index:
                continue
            raw = _normalize_cell(row[column_index])
            if not raw:
                continue
            if skip_header_row and idx == 0:
                low_hints = {h.lower() for h in _HEADER_HINTS}
                if raw in _HEADER_HINTS or raw.lower() in low_hints:
                    continue
                # 首行不像运单号（多为表头）则跳过
                if not re.fullmatch(r"[0-9A-Za-z\-]{4,50}", raw):
                    continue
            out.append(raw)
        return out
    finally:
        wb.close()


def parse_waybill_text(text: str) -> list[str]:
    """从文本框解析多个运单号（逗号、分号、换行、空白分隔）。"""
    if not (text or "").strip():
        return []
    parts = re.split(r"[\s,，;；、\n\r]+", text.strip())
    return [p.strip() for p in parts if p.strip()]


def merge_waybill_lists(*lists: list[str]) -> list[str]:
    """合并多个来源并去重，保持先后次序。"""
    seen: set[str] = set()
    ordered: list[str] = []
    for lst in lists:
        for x in lst:
            t = str(x).strip()
            if not t or t in seen:
                continue
            seen.add(t)
            ordered.append(t)
    return ordered
