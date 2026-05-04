"""excel_tracking 解析逻辑测试。"""

import io
import unittest

from excel_tracking import (
    merge_waybill_lists,
    parse_waybill_text,
    read_waybills_from_xlsx,
    read_waybills_with_dhl_extras_from_xlsx,
)


class TestParseWaybillText(unittest.TestCase):
    def test_split(self):
        self.assertEqual(
            parse_waybill_text("3112411665, 1113656224\n9294612412"),
            ["3112411665", "1113656224", "9294612412"],
        )


class TestMerge(unittest.TestCase):
    def test_dedupe_order(self):
        self.assertEqual(
            merge_waybill_lists(["1", "2"], ["2", "3"]),
            ["1", "2", "3"],
        )


class TestReadXlsx(unittest.TestCase):
    def test_column_c(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl 未安装")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["币种", "参考号", "转单号"])
        ws.append(["USD", "R1", "3112411665"])
        ws.append(["EUR", "R2", "1113656224"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ids = read_waybills_from_xlsx(buf.read(), column_index=2)
        self.assertEqual(ids, ["3112411665", "1113656224"])

    def test_dhl_extras_columns(self):
        """F/H/K 列随转单号读出（列索引与 DHL 主表一致）。"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl 未安装")

        wb = openpyxl.Workbook()
        ws = wb.active
        # A..K 表头 + 一行数据（C 转单号，F 重量，H 数量，K 价值）
        ws.append(["A"] * 11)
        ws[1][2].value = "转单号"
        ws.append(
            [
                "USD",
                "REF",
                "3112411665",
                "",
                "",
                2.5,
                "品",
                3,
                "US",
                "12345",
                99.5,
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ids, extras = read_waybills_with_dhl_extras_from_xlsx(buf.read(), waybill_column_index=2)
        self.assertEqual(ids, ["3112411665"])
        self.assertEqual(
            extras["3112411665"],
            {"重量": "2.5", "产品描述": "品", "数量": "3", "价值": "99.5"},
        )


if __name__ == "__main__":
    unittest.main()
